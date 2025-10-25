import argparse
import csv
import os
from datetime import datetime, date
from typing import Optional, List, Literal

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from .db import SessionLocal
from . import models
from .services.metrics import parse_art_interval


COLUMNS = [
    "Team Name",
    "Chatter name",
    "Shift Hours (Scheduled)",
    "Shift Hours (Actual)",
    "Date",
    "Day",
    "Sales",
    "Sold",
    "Retention",
    "Unlock",
    "Total",
    "SPH",
    "ART",
    "Golden ratio",
    "Hinge top up",
    "Tricks TSF",
    "Remarks/ Note",
]

# New simplified template columns (Excel sheet: 'Import' or CSV headers)
# Updated: remove Ranking (auto-calculated) and add Total Sales right after Chatter
NEW_COLUMNS = [
    "Chatter",
    "Total Sales",
    "Start Date",
    "End Date",
    "Worked Hrs",
    "SPH",
    "ART",
    "GR",
    "UR",
    "Shift",
]

MAX_DUPLICATE_PREVIEW = 25


def upsert_team(db: Session, name: Optional[str]) -> Optional[int]:
    if not name:
        return None
    t = db.query(models.Team).filter(models.Team.name == name).first()
    if not t:
        t = models.Team(name=name)
        db.add(t)
        db.flush()
    return t.id


def upsert_chatter(db: Session, name: str, team_id: Optional[int]) -> int:
    c = db.query(models.Chatter).filter(models.Chatter.name == name).first()
    if not c:
        c = models.Chatter(name=name, team_id=team_id, is_active=True)
        db.add(c)
        db.flush()
    else:
        if team_id and c.team_id != team_id:
            c.team_id = team_id
    return c.id


def upsert_performance(
    db: Session,
    chatter_id: int,
    team_id: Optional[int],
    dt: date,
    row: dict,
) -> Literal["created", "updated", "skipped"]:
    # Interpret incoming 'Unlock' value as either a count or a rate (fraction or percent).
    raw_unlock = row.get("Unlock")
    unlock_count_val = _to_int(raw_unlock)
    unlock_ratio_val = None
    try:
        urf = _to_float(raw_unlock)
        if urf is not None:
            # If user supplied a percentage like '5' treat as 5% -> 0.05
            # If user supplied a fraction like 0.05, keep as-is.
            if urf > 1 and urf <= 100:
                unlock_ratio_val = urf / 100.0
            elif 0 <= urf <= 1:
                unlock_ratio_val = urf
            else:
                # values outside expected ranges are ignored for ratio
                unlock_ratio_val = None
    except Exception:
        unlock_ratio_val = None

    payload = {
        "sales_amount": _to_float(row.get("Sales")),
        "sold_count": _to_int(row.get("Sold")),
        "retention_count": _to_int(row.get("Retention")),
        "unlock_count": unlock_count_val,
        "unlock_ratio": unlock_ratio_val,
        "total_sales": _to_float(row.get("Total")),
        "sph": _to_float(row.get("SPH")),
        "golden_ratio": _to_float(row.get("Golden ratio")),
        "hinge_top_up": _to_float(row.get("Hinge top up")),
        "tricks_tsf": _to_float(row.get("Tricks TSF")),
    }
    art_interval = parse_art_interval(row.get("ART"))

    existing = (
        db.query(models.PerformanceDaily)
        .filter(models.PerformanceDaily.chatter_id == chatter_id, models.PerformanceDaily.shift_date == dt)
        .first()
    )
    if existing:
        updated = False
        if team_id and existing.team_id in (None, 0):
            existing.team_id = team_id
            updated = True

        for field, value in payload.items():
            # allow setting unlock_ratio even if unlock_count exists; respect provided fields
            if value is not None:
                # special-case: only overwrite if target field is None to avoid clobbering
                if getattr(existing, field) is None:
                    setattr(existing, field, value)
                    updated = True

        if art_interval is not None and existing.art_interval is None:
            existing.art_interval = art_interval  # type: ignore[assignment]
            updated = True

        return "updated" if updated else "skipped"

    # Create new performance record; include unlock_ratio if provided
    p_kwargs = {k: v for k, v in payload.items() if v is not None}
    p = models.PerformanceDaily(
        chatter_id=chatter_id,
        team_id=team_id,
        shift_date=dt,
        **p_kwargs,
    )
    p.art_interval = art_interval  # type: ignore[assignment]
    db.add(p)
    return "created"


def upsert_shift(
    db: Session,
    chatter_id: int,
    team_id: Optional[int],
    dt: date,
    row: dict,
) -> Literal["created", "updated", "skipped"]:
    scheduled = _to_float(row.get("Shift Hours (Scheduled)"))
    actual = _to_float(row.get("Shift Hours (Actual)"))
    remarks = row.get("Remarks/ Note") or None
    shift_label = (row.get("Day") or row.get("Shift")) or None

    if scheduled is None and actual is None and remarks is None:
        # Still persist shift label if provided so downstream dashboards can reference it
        if not shift_label:
            return "skipped"

    existing = (
        db.query(models.Shift)
        .filter(models.Shift.chatter_id == chatter_id, models.Shift.shift_date == dt)
        .first()
    )
    if existing:
        updated = False
        if team_id and existing.team_id in (None, 0):
            existing.team_id = team_id
            updated = True
        if scheduled is not None and existing.scheduled_hours is None:
            existing.scheduled_hours = scheduled
            updated = True
        if actual is not None and existing.actual_hours is None:
            existing.actual_hours = actual
            updated = True
        if remarks and (existing.remarks is None or existing.remarks.strip() == ""):
            existing.remarks = remarks
            updated = True
        if shift_label and not existing.shift_day:
            existing.shift_day = shift_label
            updated = True

        return "updated" if updated else "skipped"

    sh = models.Shift(
        chatter_id=chatter_id,
        team_id=team_id,
        shift_date=dt,
        shift_day=shift_label,
        scheduled_hours=scheduled,
        actual_hours=actual,
        remarks=remarks,
    )
    db.add(sh)
    return "created"


def _to_float(v):
    try:
        if v is None or v == "":
            return None
        return float(v)
    except Exception:
        return None


def _to_int(v):
    try:
        if v is None or v == "":
            return None
        return int(v)
    except Exception:
        return None


def import_from_file(db: Session, path: str) -> dict:
    """Import data from file and return statistics.

    Supported formats:
    - Legacy Excel 'Sheet3' or CSV with COLUMNS
    - New Excel 'Import' sheet or CSV with NEW_COLUMNS (mm/dd/yyyy dates only)
    """
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xls"):
        # Detect which template is present
        wb = load_workbook(path, data_only=True)
        # Prefer new template if 'Import' sheet present and matches headers
        if 'Import' in wb.sheetnames:
            ws = wb['Import']
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            if headers and all(h in headers for h in NEW_COLUMNS):
                return _import_excel_new(db, wb, ws)
        # Fallback to legacy 'Sheet3'
        return _import_excel_legacy(db, wb)
    if ext == ".csv":
        # Peek headers to decide format
        with open(path, 'r', encoding='utf-8-sig', newline='') as f:
            reader = csv.reader(f)
            headers = next(reader, None) or []
        # Normalize BOM and whitespace
        headers = [h.strip() for h in headers]
        if headers and all(h in headers for h in NEW_COLUMNS):
            return _import_csv_new(db, path)
        return _import_csv_legacy(db, path)
    raise RuntimeError("Unsupported file type. Allowed: .xlsx, .xls, .csv")


def _import_excel_legacy(db: Session, wb) -> dict:
    if 'Sheet3' not in wb.sheetnames:
        raise RuntimeError("Sheet3 not found")
    ws = wb['Sheet3']

    # header
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    idx_map = {h: i for i, h in enumerate(headers)}

    missing = [h for h in COLUMNS if h not in idx_map]
    if missing:
        raise RuntimeError(f"Missing columns: {missing}")

    teams_created = 0
    chatters_created = 0
    perf_records = 0
    perf_updates = 0
    shift_records = 0
    shift_updates = 0
    duplicate_rows: List[dict] = []

    for row in ws.iter_rows(min_row=2):
        values = {h: row[idx_map[h]].value for h in COLUMNS}
        # Normalize to expected types
        _team = values.get("Team Name")
        team_name: Optional[str] = (str(_team).strip() if _team not in (None, "") else None)
        _ch_name = values.get("Chatter name")
        chatter_name: Optional[str] = (str(_ch_name).strip() if _ch_name not in (None, "") else None)
        date_val = values.get("Date")
        if not chatter_name or not date_val:
            continue
        # Excel date may be datetime/date or string
        if isinstance(date_val, datetime):
            dt = date_val.date()
        else:
            dt = _parse_date_flexible(str(date_val))

        # Track new creations
        existing_team_count = db.query(models.Team).count()
        team_id = upsert_team(db, team_name)
        if db.query(models.Team).count() > existing_team_count:
            teams_created += 1

        existing_chatter_count = db.query(models.Chatter).count()
        chatter_id = upsert_chatter(db, chatter_name, team_id)
        if db.query(models.Chatter).count() > existing_chatter_count:
            chatters_created += 1

        perf_status = upsert_performance(db, chatter_id, team_id, dt, values)
        if perf_status == "created":
            perf_records += 1
        elif perf_status == "updated":
            perf_updates += 1

        scheduled_present = _to_float(values.get("Shift Hours (Scheduled)")) is not None
        actual_present = _to_float(values.get("Shift Hours (Actual)")) is not None
        shift_attempted = scheduled_present or actual_present
        shift_status: Literal["created", "updated", "skipped"] = "skipped"
        if shift_attempted:
            shift_status = upsert_shift(db, chatter_id, team_id, dt, values)
            if shift_status == "created":
                shift_records += 1
            elif shift_status == "updated":
                shift_updates += 1

        if perf_status == "skipped" or (shift_attempted and shift_status == "skipped"):
            reasons = []
            if perf_status == "skipped":
                reasons.append("performance already imported")
            if shift_attempted and shift_status == "skipped":
                reasons.append("shift already imported")
            duplicate_rows.append({
                "row": row[0].row if row else None,
                "chatter": chatter_name,
                "date": dt.isoformat(),
                "details": ", ".join(reasons),
            })

    db.commit()
    return {
        "teams_created": teams_created,
        "chatters_created": chatters_created,
        "performance_records": perf_records,
        "performance_updates": perf_updates,
        "shift_records": shift_records,
        "shift_updates": shift_updates,
        "rows_skipped": len(duplicate_rows),
        "skipped_samples": duplicate_rows[:MAX_DUPLICATE_PREVIEW],
    }


def _import_csv_legacy(db: Session, path: str) -> dict:
    """Import CSV with the same columns as COLUMNS (Sheet3 format)."""
    teams_created = 0
    chatters_created = 0
    perf_records = 0
    perf_updates = 0
    shift_records = 0
    shift_updates = 0
    duplicate_rows: List[dict] = []

    with open(path, 'r', encoding='utf-8-sig', newline='') as f:
        reader = csv.DictReader(f)
        headers = reader.fieldnames or []
        missing = [h for h in COLUMNS if h not in headers]
        if missing:
            raise RuntimeError(f"Missing columns: {missing}")

        for row_idx, values in enumerate(reader, start=2):
            team_name = values.get("Team Name")
            chatter_name = values.get("Chatter name")
            date_str = values.get("Date")
            if not chatter_name or not date_str:
                continue

            try:
                dt = _parse_date_flexible(date_str)
            except Exception:
                # skip rows with unparseable date
                continue

            existing_team_count = db.query(models.Team).count()
            team_id = upsert_team(db, team_name)
            if db.query(models.Team).count() > existing_team_count:
                teams_created += 1

            existing_chatter_count = db.query(models.Chatter).count()
            chatter_id = upsert_chatter(db, chatter_name, team_id)
            if db.query(models.Chatter).count() > existing_chatter_count:
                chatters_created += 1

            perf_status = upsert_performance(db, chatter_id, team_id, dt, values)
            if perf_status == "created":
                perf_records += 1
            elif perf_status == "updated":
                perf_updates += 1

            scheduled_present = _to_float(values.get("Shift Hours (Scheduled)")) is not None
            actual_present = _to_float(values.get("Shift Hours (Actual)")) is not None
            shift_attempted = scheduled_present or actual_present
            shift_status: Literal["created", "updated", "skipped"] = "skipped"
            if shift_attempted:
                shift_status = upsert_shift(db, chatter_id, team_id, dt, values)
                if shift_status == "created":
                    shift_records += 1
                elif shift_status == "updated":
                    shift_updates += 1

            if perf_status == "skipped" or (shift_attempted and shift_status == "skipped"):
                reasons = []
                if perf_status == "skipped":
                    reasons.append("performance already imported")
                if shift_attempted and shift_status == "skipped":
                    reasons.append("shift already imported")
                duplicate_rows.append({
                    "row": row_idx,
                    "chatter": chatter_name,
                    "date": dt.isoformat(),
                    "details": ", ".join(reasons),
                })

    db.commit()
    return {
        "teams_created": teams_created,
        "chatters_created": chatters_created,
        "performance_records": perf_records,
        "performance_updates": perf_updates,
        "shift_records": shift_records,
        "shift_updates": shift_updates,
        "rows_skipped": len(duplicate_rows),
        "skipped_samples": duplicate_rows[:MAX_DUPLICATE_PREVIEW],
    }


def _import_excel_new(db: Session, wb, ws) -> dict:
    # Validate headers
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    idx_map = {h: i for i, h in enumerate(headers)}
    missing = [h for h in NEW_COLUMNS if h not in idx_map]
    if missing:
        raise RuntimeError(f"Missing columns: {missing}")

    chatters_created = 0
    perf_records = 0
    perf_updates = 0
    shift_records = 0
    shift_updates = 0
    duplicate_rows: List[dict] = []

    for row in ws.iter_rows(min_row=2):
        val = {h: row[idx_map[h]].value for h in NEW_COLUMNS}
        chatter_name = (str(val.get('Chatter')).strip() if val.get('Chatter') not in (None, '') else None)
        start_date = val.get('Start Date')
        total_sales = _to_float(val.get('Total Sales'))
        worked_hrs = _to_float(val.get('Worked Hrs'))
        sph = _to_float(val.get('SPH'))
        art = val.get('ART')
        gr = _to_float(val.get('GR'))
        ur = _to_float(val.get('UR'))
        shift_label = (val.get('Shift') or '').strip() or None

        if not chatter_name or not start_date:
            continue

        # Enforce mm/dd/yyyy only
        if isinstance(start_date, (datetime, date)):
            dt = start_date if isinstance(start_date, date) else start_date.date()
        else:
            dt = _parse_mmddyyyy_only(str(start_date))

        existing_chatter_count = db.query(models.Chatter).count()
        chatter_id = upsert_chatter(db, chatter_name, None)
        if db.query(models.Chatter).count() > existing_chatter_count:
            chatters_created += 1

        # Build legacy-like row mapping for reuse
        sales_amount = total_sales if total_sales is not None else (
            (worked_hrs * sph) if (worked_hrs is not None and sph is not None) else None
        )
        values = {
            "Sales": sales_amount,
            "Sold": None,
            "Retention": None,
            "Unlock": ur,  # treat UR as unlock rate; stored in unlock_count (best-effort)
            "Total": sales_amount,
            "SPH": sph,
            "ART": art,
            "Golden ratio": gr,
            "Hinge top up": None,
            "Tricks TSF": None,
            "Day": None,
            "Shift Hours (Actual)": worked_hrs,
            "Shift Hours (Scheduled)": None,
            "Remarks/ Note": None,
            "Shift": shift_label,
            "Day": shift_label,
        }

        perf_status = upsert_performance(db, chatter_id, None, dt, values)
        if perf_status == "created":
            perf_records += 1
        elif perf_status == "updated":
            perf_updates += 1

        shift_attempted = (worked_hrs is not None) or (shift_label is not None)
        shift_status: Literal["created", "updated", "skipped"] = "skipped"
        if shift_attempted:
            shift_status = upsert_shift(db, chatter_id, None, dt, values)
            if shift_status == "created":
                shift_records += 1
            elif shift_status == "updated":
                shift_updates += 1

        if perf_status == "skipped" or (shift_attempted and shift_status == "skipped"):
            duplicate_rows.append({
                "row": row[0].row if row else None,
                "chatter": chatter_name,
                "date": dt.isoformat(),
                "details": ", ".join(filter(None, [
                    None if perf_status != "skipped" else "performance already imported",
                    None if (not shift_attempted or shift_status != "skipped") else "shift already imported",
                ])),
            })

    db.commit()
    return {
        "teams_created": 0,
        "chatters_created": chatters_created,
        "performance_records": perf_records,
        "performance_updates": perf_updates,
        "shift_records": shift_records,
        "shift_updates": shift_updates,
        "rows_skipped": len(duplicate_rows),
        "skipped_samples": duplicate_rows[:MAX_DUPLICATE_PREVIEW],
    }


def _import_csv_new(db: Session, path: str) -> dict:
    chatters_created = 0
    perf_records = 0
    perf_updates = 0
    shift_records = 0
    shift_updates = 0
    duplicate_rows: List[dict] = []

    with open(path, 'r', encoding='utf-8-sig', newline='') as f:
        reader = csv.DictReader(f)
        headers = reader.fieldnames or []
        missing = [h for h in NEW_COLUMNS if h not in headers]
        if missing:
            raise RuntimeError(f"Missing columns: {missing}")

        for row_idx, val in enumerate(reader, start=2):
            chatter_name = (val.get('Chatter') or '').strip() or None
            start_date = val.get('Start Date')
            total_sales = _to_float(val.get('Total Sales'))
            worked_hrs = _to_float(val.get('Worked Hrs'))
            sph = _to_float(val.get('SPH'))
            art = val.get('ART')
            gr = _to_float(val.get('GR'))
            ur = _to_float(val.get('UR'))
            shift_label = (val.get('Shift') or '').strip() or None
            if not chatter_name or not start_date:
                continue

            try:
                dt = _parse_mmddyyyy_only(str(start_date))
            except Exception:
                # skip rows with unparseable date (strict mm/dd/yyyy)
                continue

            existing_chatter_count = db.query(models.Chatter).count()
            chatter_id = upsert_chatter(db, chatter_name, None)
            if db.query(models.Chatter).count() > existing_chatter_count:
                chatters_created += 1

            sales_amount = total_sales if total_sales is not None else (
                (worked_hrs * sph) if (worked_hrs is not None and sph is not None) else None
            )
            values = {
                "Sales": sales_amount,
                "Sold": None,
                "Retention": None,
                "Unlock": ur,
                "Total": sales_amount,
                "SPH": sph,
                "ART": art,
                "Golden ratio": gr,
                "Hinge top up": None,
                "Tricks TSF": None,
                "Day": None,
                "Shift Hours (Actual)": worked_hrs,
                "Shift Hours (Scheduled)": None,
                "Remarks/ Note": None,
                "Shift": shift_label,
                "Day": shift_label,
            }

            perf_status = upsert_performance(db, chatter_id, None, dt, values)
            if perf_status == "created":
                perf_records += 1
            elif perf_status == "updated":
                perf_updates += 1

            shift_attempted = (worked_hrs is not None) or (shift_label is not None)
            shift_status: Literal["created", "updated", "skipped"] = "skipped"
            if shift_attempted:
                shift_status = upsert_shift(db, chatter_id, None, dt, values)
                if shift_status == "created":
                    shift_records += 1
                elif shift_status == "updated":
                    shift_updates += 1

            if perf_status == "skipped" or (shift_attempted and shift_status == "skipped"):
                duplicate_rows.append({
                    "row": row_idx,
                    "chatter": chatter_name,
                    "date": dt.isoformat(),
                    "details": ", ".join(filter(None, [
                        None if perf_status != "skipped" else "performance already imported",
                        None if (not shift_attempted or shift_status != "skipped") else "shift already imported",
                    ])),
                })

    db.commit()
    return {
        "teams_created": 0,
        "chatters_created": chatters_created,
        "performance_records": perf_records,
        "performance_updates": perf_updates,
        "shift_records": shift_records,
        "shift_updates": shift_updates,
        "rows_skipped": len(duplicate_rows),
        "skipped_samples": duplicate_rows[:MAX_DUPLICATE_PREVIEW],
    }


def _parse_date_flexible(value: str) -> date:
    """Parse date with preference for mm/dd/yyyy or ISO yyyy-mm-dd.

    Rationale:
    - Our data entry and templates standardize on MM/DD/YYYY.
    - Ambiguous day-first formats (DD/MM/YYYY) can corrupt months when day<=12.
      To avoid this, we no longer auto-accept DD/MM/YYYY for slash-delimited dates.
    """
    s = value.strip()
    # Explicit ISO (yyyy-mm-dd)
    try:
        # Strict ISO first
        return datetime.strptime(s, "%Y-%m-%d").date()
    except Exception:
        pass
    # Strict mm/dd/yyyy for slash dates
    try:
        return datetime.strptime(s, "%m/%d/%Y").date()
    except Exception:
        pass
    # As a last resort, attempt Python's ISO-like parser (handles yyyy-mm-ddThh:mm:ss)
    try:
        return datetime.fromisoformat(s).date()
    except Exception:
        pass
    raise ValueError(f"Unrecognized date format (expected MM/DD/YYYY or YYYY-MM-DD): {value}")


def _parse_mmddyyyy_only(value: str) -> date:
    """Parse date strictly in mm/dd/yyyy format."""
    return datetime.strptime(value.strip(), "%m/%d/%Y").date()


def main(path: str):
    db = SessionLocal()
    try:
        stats = import_from_file(db, path)
        print(f"Import complete: {stats}")
    finally:
        db.close()


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Import Excel Sheet3 for of-chatters")
    parser.add_argument("--path", required=True, help="Path to Excel file")
    args = parser.parse_args()
    main(args.path)
